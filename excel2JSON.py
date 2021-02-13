from openpyxl import load_workbook
import json

class Excel2JSON:
    def __init__(self):
        workbook = load_workbook(filename="UserFile.xlsx")
        sheet = workbook.active
        self.rows = []
        for row in sheet.iter_rows(values_only=True):
            self.rows.append(row)
        self.extractHeader()
    
    def extractHeader(self):
        self.head = self.rows[0]
        self.rows = self.rows[1:]
        self.createUserList()

    def createUserList(self):
        count = len(self.head)
        userList = []
        for row in self.rows:
            if row.__contains__(None):
                continue
            userObject={}
            userObject["userName"] = row[0]
            userObject["userAttribute"]=[]
            for i in range(1,count-1):
                name = self.head[i]
                value = row[i]
                if value == "=TRUE()":
                    value = "true"
                elif value == "=FALSE()":
                    value = "false"
                if name.__contains__("_verified"):
                    value = str(value).lower()                    
                userObject["userAttribute"].append(
                    {
                        "Name": name,
                        "Value": value
                    }
                )
            userObject["group"] = row[-1].split(',')
            userList.append(userObject)
        self.writeToJSON(userList=userList)

    def writeToJSON(self,userList):
        with open("data.json", "w+") as file:
            json.dump(userList,fp=file)